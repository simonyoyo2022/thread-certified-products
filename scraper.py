#!/usr/bin/env python3
"""
Thread Group Certified Products Scraper — Fixed Version
Uses correct HTML patterns discovered via debug inspection.

Strategy:
  Phase 1 (129 requests): Scan each company → get products + logo URLs
  Phase 2 (9 requests):   Scan each Device Type globally → map (logo, name) → device_type
  Phase 3 (30 requests):  Scan each Sub Category globally → map (logo, name) → sub_categories
  Final:                  Join all data and export
"""

import urllib.request
import urllib.parse
import re
import json
import http.cookiejar
import time
import os
import logging
from datetime import datetime
from collections import defaultdict

logger = logging.getLogger(__name__)

BASE_URL = "https://threadgroup.org/Certified-Products"
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
PRODUCTS_FILE = os.path.join(DATA_DIR, "products.json")
LOG_FILE = os.path.join(DATA_DIR, "scrape_log.json")

# Device Types in the filter panel
DEVICE_TYPES = [
    (0, "Automation control"),
    (1, "Chipset"),
    (2, "HVAC"),
    (3, "Infrastructure"),
    (4, "Lighting"),
    (5, "Module"),
    (6, "Safety"),
    (7, "Sensor"),
    (8, "Window covering"),
]

# Sub Categories in the filter panel
SUB_CATEGORIES = [
    (0, "Agricultural"), (1, "Air"), (2, "Air conditioning"),
    (3, "Blind"), (4, "Bulb"), (5, "Contact"),
    (6, "Curtain"), (7, "Door lock"), (8, "Gas alarm"),
    (9, "Gateway"), (10, "Hub"), (11, "Irrigation"),
    (12, "Light"), (13, "Lighting"), (14, "Lightstrip"),
    (15, "Media streamer"), (16, "Presence"), (17, "Pump"),
    (18, "Shutter"), (19, "Smart display"), (20, "Smart plug"),
    (21, "Smart speaker"), (22, "Smoke alarm"), (23, "Switch"),
    (24, "Temperature"), (25, "Water"), (26, "Weather"),
    (27, "Wi-Fi Access Point"), (28, "Window lock"),
]

_progress_callback = None
_cancel_flag = False


def set_progress_callback(fn):
    global _progress_callback
    _progress_callback = fn


def cancel_scrape():
    global _cancel_flag
    _cancel_flag = True


def report_progress(step, total, msg):
    if _progress_callback:
        _progress_callback(step, total, msg)
    logger.info(f"[{step}/{total}] {msg}")


def create_opener():
    jar = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(jar))
    opener.addheaders = [
        ('User-Agent', 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'),
        ('Accept', 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'),
        ('Accept-Language', 'en-US,en;q=0.5'),
        ('Connection', 'keep-alive'),
    ]
    return opener


def extract_viewstate(html):
    vs  = re.search(r'id="__VIEWSTATE"[^>]*value="([^"]*)"', html)
    vsg = re.search(r'id="__VIEWSTATEGENERATOR"[^>]*value="([^"]*)"', html)
    ev  = re.search(r'id="__EVENTVALIDATION"[^>]*value="([^"]*)"', html)
    return {
        '__VIEWSTATE':          vs.group(1)  if vs  else '',
        '__VIEWSTATEGENERATOR': vsg.group(1) if vsg else 'CA0B0334',
        '__EVENTVALIDATION':    ev.group(1)  if ev  else '',
    }


def update_vs_from_ajax(vs, ajax_html):
    """Update ViewState/EventValidation from AJAX delta response."""
    vs_m  = re.search(r'\d+\|hiddenField\|__VIEWSTATE\|([^|]+)', ajax_html)
    ev_m  = re.search(r'\d+\|hiddenField\|__EVENTVALIDATION\|([^|]+)', ajax_html)
    vsg_m = re.search(r'\d+\|hiddenField\|__VIEWSTATEGENERATOR\|([^|]+)', ajax_html)
    if vs_m:  vs['__VIEWSTATE'] = vs_m.group(1)
    if ev_m:  vs['__EVENTVALIDATION'] = ev_m.group(1)
    if vsg_m: vs['__VIEWSTATEGENERATOR'] = vsg_m.group(1)
    return vs


def extract_companies_from_html(html):
    """
    Extract all company names using input+label pattern.
    HTML: <input name="...chkCompanyName$N" value="V" /><label>Name</label>
    """
    pattern = re.compile(
        r'<input[^>]*name="dnn\$ctr1641\$Default\$ctl00\$chkCompanyName\$(\d+)"'
        r'[^>]*value="(\d+)"[^>]*/>'
        r'<label[^>]*>([^<]+)</label>',
        re.IGNORECASE
    )
    companies = []
    for m in pattern.finditer(html):
        name = m.group(3).strip()
        name = re.sub(r'&amp;', '&', name)
        name = re.sub(r'&#\d+;', '', name).strip()
        companies.append({'index': int(m.group(1)), 'value': int(m.group(2)), 'name': name})
    return companies


def extract_products_from_html(html):
    """
    Extract products by splitting on <div id="prod-sec1">.
    Each split part contains: logo URL, badge type, then product name in <h1>.
    """
    tag_re = re.compile(r'<[^>]+>')
    products = []

    parts = re.split(r'<div id="prod-sec1">', html)

    for part in parts[1:]:  # Skip content before first product
        # Company logo URL
        logo_m = re.search(
            r'src="(https://membershipcore\.inventures\.com//[^"]+MemberLogo/[^"]+)"',
            part)
        logo_url = logo_m.group(1) if logo_m else ''

        # Product type badge
        bot = bool(re.search(r'TG_BOT_g\.png', part, re.IGNORECASE))
        prod_type = "Built on Thread" if bot else "Thread Certified Component"

        # Product name from <h1> (in prod-sec2)
        h1_m = re.search(r'<h1[^>]*>(.*?)</h1>', part, re.DOTALL)
        if not h1_m:
            continue
        name = tag_re.sub('', h1_m.group(1)).strip()
        if not name or name.lower() == 'filters':
            continue

        products.append({
            'name': name,
            'logo_url': logo_url,
            'product_type': prod_type,
        })

    return products


def make_ajax_request(opener, vs, event_target, extra_fields=None, retries=3):
    """Send an ASP.NET ScriptManager AJAX postback and return response HTML."""
    post = {
        '__EVENTTARGET':        event_target,
        '__EVENTARGUMENT':      '',
        '__LASTFOCUS':          '',
        '__VIEWSTATE':          vs['__VIEWSTATE'],
        '__VIEWSTATEGENERATOR': vs['__VIEWSTATEGENERATOR'],
        '__VIEWSTATEENCRYPTED': '',
        '__EVENTVALIDATION':    vs['__EVENTVALIDATION'],
        'ScriptManager': (
            f'dnn$ctr1641$Default$ctl00$UpdatePanel1|{event_target}'
        ),
        '__ASYNCPOST': 'true',
        'dnn$ctr1641$Default$ctl00$ddlPageSize': '20',
    }
    if extra_fields:
        post.update(extra_fields)

    for attempt in range(retries):
        try:
            req = urllib.request.Request(
                BASE_URL,
                data=urllib.parse.urlencode(post).encode('utf-8'))
            req.add_header('User-Agent',
                           'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)')
            req.add_header('Content-Type',
                           'application/x-www-form-urlencoded; charset=UTF-8')
            req.add_header('X-MicrosoftAjax', 'Delta=true')
            req.add_header('X-Requested-With', 'XMLHttpRequest')
            req.add_header('Referer', BASE_URL)
            req.add_header('Origin', 'https://threadgroup.org')
            with opener.open(req, timeout=30) as resp:
                return resp.read().decode('utf-8', errors='replace')
        except Exception as e:
            if attempt < retries - 1:
                logger.warning(f"Request failed ({attempt+1}/{retries}): {e}. Retrying...")
                time.sleep(2 ** attempt)
            else:
                raise


def logo_key(logo_url):
    """Extract a short identifier from logo URL for matching."""
    # e.g. "1871578067Core_Tuya logo.png" → "1871578067"
    m = re.search(r'MemberLogo/(\d+)', logo_url)
    return m.group(1) if m else logo_url


class ThreadGroupScraper:
    def __init__(self):
        self.opener = None
        self.vs = {}
        self.companies = []       # [{index, value, name}]
        # Phase 1 results: company_name → {logo_key, products: [(name, prod_type)]}
        self.company_data = {}
        # logo_key → company_name (built during phase 1)
        self.logo_to_company = {}
        # (logo_key, product_name) → device_type
        self.device_type_map = {}
        # (logo_key, product_name) → [sub_categories]
        self.sub_cat_map = defaultdict(list)

    # ── Init ──────────────────────────────────────────────
    def init(self):
        global _cancel_flag
        _cancel_flag = False

        self.opener = create_opener()
        report_progress(0, 100, "Loading Thread Group Certified Products page...")

        req = urllib.request.Request(BASE_URL)
        req.add_header('User-Agent',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)')
        with self.opener.open(req, timeout=30) as resp:
            html = resp.read().decode('utf-8', errors='replace')

        self.vs = extract_viewstate(html)
        self.companies = extract_companies_from_html(html)
        logger.info(f"Found {len(self.companies)} companies.")
        report_progress(1, 100, f"Found {len(self.companies)} companies")
        return html

    # ── Helper ────────────────────────────────────────────
    def _ajax(self, event_target, extra=None):
        time.sleep(0.5)
        html = make_ajax_request(self.opener, self.vs, event_target, extra)
        update_vs_from_ajax(self.vs, html)
        return html

    def _uncheck(self, event_target):
        """Uncheck a filter (send the same target without the value)."""
        self._ajax(event_target)

    # ── Phase 1: Company scan ─────────────────────────────
    def phase1_scan_companies(self):
        """Scan each company to get product list and logo URL."""
        total = len(self.companies)
        report_progress(5, 100, f"Phase 1: Scanning {total} companies...")

        for i, company in enumerate(self.companies):
            if _cancel_flag:
                return

            prog = 5 + int(i * 45 / total)
            report_progress(prog, 100,
                            f"  [{i+1}/{total}] {company['name']}")

            try:
                # Filter by this company
                html = self._ajax(
                    f'dnn$ctr1641$Default$ctl00$chkCompanyName${company["index"]}',
                    {f'dnn$ctr1641$Default$ctl00$chkCompanyName${company["index"]}':
                     str(company['value'])}
                )

                products = extract_products_from_html(html)

                # Store company data
                co_name = company['name']
                self.company_data[co_name] = {
                    'products': [(p['name'], p['product_type']) for p in products]
                }

                # Build logo → company mapping
                for p in products:
                    lk = logo_key(p['logo_url'])
                    if lk:
                        self.logo_to_company[lk] = co_name

                logger.info(
                    f"  {co_name}: {len(products)} products")

                # Uncheck company
                self._uncheck(
                    f'dnn$ctr1641$Default$ctl00$chkCompanyName${company["index"]}')

            except Exception as e:
                logger.warning(f"Error scanning '{company['name']}': {e}")

    # ── Phase 2: Device Type global scan ──────────────────
    def phase2_scan_device_types(self):
        """Globally scan each Device Type to map products."""
        report_progress(52, 100,
                        "Phase 2: Scanning Device Types globally...")

        for dt_idx, dt_name in DEVICE_TYPES:
            if _cancel_flag:
                return

            prog = 52 + int(dt_idx * 10 / len(DEVICE_TYPES))
            report_progress(prog, 100, f"  Device Type: {dt_name}")

            try:
                html = self._ajax(
                    f'dnn$ctr1641$Default$ctl00$chkDeviceType${dt_idx}',
                    {f'dnn$ctr1641$Default$ctl00$chkDeviceType${dt_idx}': str(dt_idx + 1)}
                )

                products = extract_products_from_html(html)
                for p in products:
                    lk = logo_key(p['logo_url'])
                    key = (lk, p['name'])
                    self.device_type_map[key] = dt_name

                logger.info(f"  DT '{dt_name}': {len(products)} products mapped")

                # Uncheck device type
                self._uncheck(
                    f'dnn$ctr1641$Default$ctl00$chkDeviceType${dt_idx}')

            except Exception as e:
                logger.warning(f"Error scanning device type '{dt_name}': {e}")

    # ── Phase 3: Sub Category global scan ─────────────────
    def phase3_scan_sub_categories(self):
        """Globally scan each Sub Category to map products."""
        report_progress(63, 100,
                        "Phase 3: Scanning Sub Categories globally...")

        for sc_idx, sc_name in SUB_CATEGORIES:
            if _cancel_flag:
                return

            prog = 63 + int(sc_idx * 30 / len(SUB_CATEGORIES))
            report_progress(prog, 100, f"  Sub Category: {sc_name}")

            try:
                html = self._ajax(
                    f'dnn$ctr1641$Default$ctl00$chkSubCategory${sc_idx}',
                    {f'dnn$ctr1641$Default$ctl00$chkSubCategory${sc_idx}': str(sc_idx + 1)}
                )

                products = extract_products_from_html(html)
                for p in products:
                    lk = logo_key(p['logo_url'])
                    key = (lk, p['name'])
                    if sc_name not in self.sub_cat_map[key]:
                        self.sub_cat_map[key].append(sc_name)

                logger.info(f"  SC '{sc_name}': {len(products)} products mapped")

                # Uncheck sub category
                self._uncheck(
                    f'dnn$ctr1641$Default$ctl00$chkSubCategory${sc_idx}')

            except Exception as e:
                logger.warning(f"Error scanning sub category '{sc_name}': {e}")

    # ── Build final results ────────────────────────────────
    def build_results(self):
        """Join phases 1, 2, 3 into the final product list."""
        results = []

        for co_name, co_info in self.company_data.items():
            # Find the logo key for this company
            co_logo_key = None
            for lk, cn in self.logo_to_company.items():
                if cn == co_name:
                    co_logo_key = lk
                    break

            for prod_name, prod_type in co_info['products']:
                # Look up device type
                dt = '-'
                if co_logo_key:
                    dt = self.device_type_map.get((co_logo_key, prod_name), '-')
                    if dt == '-':
                        # Try matching by product name only (fallback)
                        for (lk2, pn2), dt2 in self.device_type_map.items():
                            if pn2 == prod_name:
                                dt = dt2
                                break

                # Look up sub categories
                sc_list = []
                if co_logo_key:
                    sc_list = self.sub_cat_map.get((co_logo_key, prod_name), [])
                    if not sc_list:
                        for (lk2, pn2), sc2 in self.sub_cat_map.items():
                            if pn2 == prod_name:
                                sc_list = sc2
                                break

                results.append({
                    'company': co_name,
                    'product_name': prod_name,
                    'product_type': prod_type,
                    'device_type': dt,
                    'sub_category': ', '.join(sc_list) if sc_list else '-',
                })

        return sorted(results, key=lambda x: (x['company'], x['product_name']))

    # ── Save JSON ──────────────────────────────────────────
    def save_to_json(self, results):
        os.makedirs(DATA_DIR, exist_ok=True)
        output = {
            'last_updated': datetime.now().isoformat(),
            'total_products': len(results),
            'total_companies': len(set(r['company'] for r in results)),
            'products': results,
        }
        with open(PRODUCTS_FILE, 'w', encoding='utf-8') as f:
            json.dump(output, f, ensure_ascii=False, indent=2)

        log_entry = {
            'timestamp': datetime.now().isoformat(),
            'total_products': len(results),
            'total_companies': len(set(r['company'] for r in results)),
            'status': 'success',
        }
        logs = []
        if os.path.exists(LOG_FILE):
            try:
                with open(LOG_FILE) as f:
                    logs = json.load(f)
            except Exception:
                pass
        logs.append(log_entry)
        logs = logs[-50:]
        with open(LOG_FILE, 'w') as f:
            json.dump(logs, f, indent=2)

        return output

    # ── Generate Excel ─────────────────────────────────────
    def export_to_excel(self, results):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, PieChart, Reference
        from collections import Counter

        wb = openpyxl.Workbook()

        HBG = "1F3864"; HFG = "FFFFFF"; ALT = "EBF0FA"
        thin = Side(style='thin', color='CCCCCC')
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

        def hfill(c):
            return PatternFill(start_color=c, end_color=c, fill_type='solid')

        # ─ Sheet 1: All Products ─
        ws1 = wb.active
        ws1.title = "All Products"

        ws1.merge_cells('A1:F1')
        ws1['A1'].value = "Thread Group Certified Products — All Companies"
        ws1['A1'].font = Font(name='Calibri', bold=True, size=16, color=HFG)
        ws1['A1'].fill = hfill("243F60")
        ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws1.row_dimensions[1].height = 36

        ws1.merge_cells('A2:F2')
        ws1['A2'].value = (
            f"Source: threadgroup.org/Certified-Products  |  "
            f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
            f"Total: {len(results)} products, "
            f"{len(set(r['company'] for r in results))} companies"
        )
        ws1['A2'].font = Font(name='Calibri', italic=True, size=10, color="555555")
        ws1['A2'].fill = hfill("D6E4F0")
        ws1['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws1.row_dimensions[2].height = 20
        ws1.row_dimensions[3].height = 8

        headers   = ["#", "Company", "Product Name", "Product Type", "Device Type", "Sub Category"]
        col_widths = [5, 35, 45, 30, 25, 25]
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            c = ws1.cell(row=4, column=ci, value=h)
            c.font = Font(name='Calibri', bold=True, size=11, color=HFG)
            c.fill = hfill(HBG)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = bdr
            ws1.column_dimensions[get_column_letter(ci)].width = w
        ws1.row_dimensions[4].height = 28

        for ri, row in enumerate(results, 5):
            bg = ALT if ri % 2 == 1 else "FFFFFF"
            vals = [ri - 4, row['company'], row['product_name'],
                    row['product_type'], row['device_type'], row['sub_category']]
            for ci, v in enumerate(vals, 1):
                c = ws1.cell(row=ri, column=ci, value=v)
                c.fill = hfill(bg); c.border = bdr
                c.alignment = Alignment(
                    vertical='center', wrap_text=True,
                    horizontal='center' if ci in [1, 4, 5, 6] else 'left')
                c.font = Font(name='Calibri', size=10)
            ws1.row_dimensions[ri].height = 18
        ws1.freeze_panes = 'A5'

        # ─ Sheet 2: Company Summary ─
        ws2 = wb.create_sheet("Company Summary")
        ws2.merge_cells('A1:D1')
        ws2['A1'].value = "Company Summary"
        ws2['A1'].font = Font(name='Calibri', bold=True, size=14, color=HFG)
        ws2['A1'].fill = hfill("243F60")
        ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[1].height = 32

        company_counts = Counter(r['company'] for r in results)
        for ci, (h, w) in enumerate(
            zip(["#", "Company", "Products", "Device Types"],
                [5, 40, 12, 50]), 1):
            c = ws2.cell(row=3, column=ci, value=h)
            c.font = Font(bold=True, color=HFG); c.fill = hfill(HBG)
            c.alignment = Alignment(horizontal='center'); c.border = bdr
            ws2.column_dimensions[get_column_letter(ci)].width = w

        for ri, (co, cnt) in enumerate(company_counts.most_common(), 4):
            dts = set(r['device_type'] for r in results
                      if r['company'] == co and r['device_type'] != '-')
            bg = ALT if ri % 2 == 0 else "FFFFFF"
            vals2 = [ri - 3, co, cnt, ', '.join(sorted(dts))]
            for ci, v in enumerate(vals2, 1):
                c = ws2.cell(row=ri, column=ci, value=v)
                c.fill = hfill(bg); c.border = bdr
                c.font = Font(name='Calibri', size=10)
                c.alignment = Alignment(
                    horizontal='center' if ci in [1, 3] else 'left',
                    vertical='center', wrap_text=True)

        # ─ Sheet 3: Charts ─
        ws3 = wb.create_sheet("Charts")
        ws3.merge_cells('A1:H1')
        ws3['A1'].value = "Analysis Charts"
        ws3['A1'].font = Font(name='Calibri', bold=True, size=14, color=HFG)
        ws3['A1'].fill = hfill("243F60")
        ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws3.row_dimensions[1].height = 32

        dt_counts = Counter(r['device_type'] for r in results if r['device_type'] != '-')
        ws3.cell(row=3, column=1, value="Device Type").font = Font(bold=True)
        ws3.cell(row=3, column=2, value="Count").font = Font(bold=True)
        for i, (dt, cnt) in enumerate(sorted(dt_counts.items()), 1):
            ws3.cell(row=3+i, column=1, value=dt)
            ws3.cell(row=3+i, column=2, value=cnt)

        chart_dt = BarChart()
        chart_dt.type = "col"; chart_dt.title = "Products by Device Type"
        chart_dt.style = 10; chart_dt.width = 22; chart_dt.height = 14
        dr = Reference(ws3, min_col=2, max_col=2, min_row=3, max_row=3+len(dt_counts))
        cr = Reference(ws3, min_col=1, max_col=1, min_row=4, max_row=3+len(dt_counts))
        chart_dt.add_data(dr, titles_from_data=True)
        chart_dt.set_categories(cr)
        ws3.add_chart(chart_dt, "D3")

        pt_counts = Counter(r['product_type'] for r in results)
        ws3.cell(row=3, column=16, value="Product Type").font = Font(bold=True)
        ws3.cell(row=3, column=17, value="Count").font = Font(bold=True)
        for i, (pt, cnt) in enumerate(pt_counts.items(), 1):
            ws3.cell(row=3+i, column=16, value=pt)
            ws3.cell(row=3+i, column=17, value=cnt)

        chart_pt = PieChart()
        chart_pt.title = "Product Type Distribution"
        chart_pt.style = 10; chart_pt.width = 18; chart_pt.height = 14
        dr2 = Reference(ws3, min_col=17, max_col=17, min_row=3, max_row=3+len(pt_counts))
        cr2 = Reference(ws3, min_col=16, max_col=16, min_row=4, max_row=3+len(pt_counts))
        chart_pt.add_data(dr2, titles_from_data=True)
        chart_pt.set_categories(cr2)
        ws3.add_chart(chart_pt, "D20")

        top15 = company_counts.most_common(15)
        ws3.cell(row=3, column=25, value="Company").font = Font(bold=True)
        ws3.cell(row=3, column=26, value="Products").font = Font(bold=True)
        for i, (co, cnt) in enumerate(top15, 1):
            ws3.cell(row=3+i, column=25, value=co)
            ws3.cell(row=3+i, column=26, value=cnt)

        chart_co = BarChart()
        chart_co.type = "bar"; chart_co.title = "Top 15 Companies"
        chart_co.style = 10; chart_co.width = 22; chart_co.height = 16
        dr3 = Reference(ws3, min_col=26, max_col=26, min_row=3, max_row=3+15)
        cr3 = Reference(ws3, min_col=25, max_col=25, min_row=4, max_row=3+15)
        chart_co.add_data(dr3, titles_from_data=True)
        chart_co.set_categories(cr3)
        ws3.add_chart(chart_co, "D38")

        excel_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "Thread_Certified_Products_All.xlsx")
        wb.save(excel_path)
        logger.info(f"Excel saved: {excel_path}")
        return excel_path


def run_scrape(progress_callback=None):
    """Main entry point. Returns output dict or raises on error."""
    global _cancel_flag
    _cancel_flag = False

    if progress_callback:
        set_progress_callback(progress_callback)

    os.makedirs(DATA_DIR, exist_ok=True)
    scraper = ThreadGroupScraper()

    try:
        scraper.init()
        scraper.phase1_scan_companies()
        if _cancel_flag:
            return None
        scraper.phase2_scan_device_types()
        if _cancel_flag:
            return None
        scraper.phase3_scan_sub_categories()

        report_progress(95, 100, "Building final results...")
        results = scraper.build_results()

        report_progress(97, 100, "Saving JSON...")
        output = scraper.save_to_json(results)

        report_progress(99, 100, "Generating Excel...")
        excel_path = scraper.export_to_excel(results)
        output['excel_path'] = excel_path

        report_progress(100, 100,
                        f"Done! {len(results)} products from "
                        f"{output['total_companies']} companies.")
        return output

    except Exception as e:
        logger.error(f"Scrape failed: {e}", exc_info=True)
        log_entry = {
            'timestamp': datetime.now().isoformat(),
            'status': 'error',
            'error': str(e),
        }
        logs = []
        if os.path.exists(LOG_FILE):
            try:
                with open(LOG_FILE) as f:
                    logs = json.load(f)
            except Exception:
                pass
        logs.append(log_entry)
        with open(LOG_FILE, 'w') as f:
            json.dump(logs, f, indent=2)
        raise


def load_cached_data():
    if not os.path.exists(PRODUCTS_FILE):
        return None
    with open(PRODUCTS_FILE, encoding='utf-8') as f:
        return json.load(f)


if __name__ == '__main__':
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s')
    result = run_scrape()
    print(f"\n✅ Done: {result['total_products']} products, "
          f"{result['total_companies']} companies")
