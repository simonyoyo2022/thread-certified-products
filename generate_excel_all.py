#!/usr/bin/env python3
"""
Generate comprehensive Excel file for all Thread Group Certified Products.
Reads from data/products.json and creates:
  Sheet 1: All Products Data (sortable table)
  Sheet 2: Company Summary 
  Sheet 3: Charts & Analysis
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from collections import Counter, defaultdict
import json
import os
from datetime import datetime

PRODUCTS_FILE = "/Users/yoyolin/project/Thread/data/products.json"
OUTPUT_PATH   = "/Users/yoyolin/project/Thread/Thread_Certified_Products_All.xlsx"

# ── Color palette ───────────────────────────────────
C_HEADER_BG = "1F3864"
C_HEADER_FG = "FFFFFF"
C_ROW_ALT   = "EBF0FA"
C_TITLE_BG  = "243F60"
C_ACCENT1   = "2E75B6"
C_ACCENT2   = "70AD47"
C_ACCENT3   = "ED7D31"

def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

def border():
    t = Side(style='thin', color='CCCCCC')
    return Border(left=t, right=t, top=t, bottom=t)

def hdr_font(size=11):
    return Font(name='Calibri', bold=True, size=size, color=C_HEADER_FG)

def cell_font(size=10, bold=False):
    return Font(name='Calibri', size=size, bold=bold)

# ── Sheet 1: All Products ───────────────────────────
def create_products_sheet(wb, products, last_updated):
    ws = wb.active
    ws.title = "All Products"

    # Title
    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value = "Thread Group Certified Products — All Companies"
    c.font = Font(name='Calibri', bold=True, size=16, color=C_HEADER_FG)
    c.fill = fill(C_TITLE_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    # Subtitle
    ws.merge_cells('A2:G2')
    c = ws['A2']
    c.value = f"Source: threadgroup.org/Certified-Products  |  Total: {len(products)} products  |  Updated: {last_updated[:10]}"
    c.font = Font(name='Calibri', italic=True, size=10, color="555555")
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = fill("D6E4F0")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 8

    # Headers
    headers   = ["#", "Company", "Product Name", "Product Type", "Device Type", "Sub Category", "Cert. Number"]
    col_widths = [5,   30,        45,              28,             20,            20,             20]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font      = hdr_font()
        c.fill      = fill(C_HEADER_BG)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 26

    # Data rows
    for ri, p in enumerate(products, 5):
        is_alt = ri % 2 == 1
        bg     = C_ROW_ALT if is_alt else "FFFFFF"
        values = [
            ri - 4,
            p.get('company', ''),
            p.get('product_name', ''),
            p.get('product_type', ''),
            p.get('device_type', '-'),
            p.get('sub_category', '-'),
            p.get('cert_number', ''),
        ]
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.fill      = fill(bg)
            c.border    = border()
            c.font      = cell_font()
            c.alignment = Alignment(
                horizontal='center' if ci in [1, 4, 5, 6] else 'left',
                vertical='center', wrap_text=True
            )
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f"A4:G{len(products)+4}"

    return ws


# ── Sheet 2: Company Summary ────────────────────────
def create_company_sheet(wb, products):
    ws = wb.create_sheet("Company Summary")

    # Title
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = "Company Summary — Products by Device Type & Category"
    c.font  = Font(name='Calibri', bold=True, size=14, color=C_HEADER_FG)
    c.fill  = fill(C_TITLE_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 8

    headers   = ["Company", "Total Products", "Device Types", "Sub Categories", "Product Types"]
    col_widths = [35,         16,               35,             40,              30]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font      = hdr_font()
        c.fill      = fill(C_HEADER_BG)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 26

    # Group by company
    companies = defaultdict(lambda: {'total': 0, 'dts': set(), 'scs': set(), 'pts': set()})
    for p in products:
        co = p.get('company', 'Unknown')
        companies[co]['total'] += 1
        dt = p.get('device_type', '-')
        if dt and dt != '-':
            companies[co]['dts'].add(dt)
        sc = p.get('sub_category', '-')
        if sc and sc != '-':
            for s in sc.split(', '):
                companies[co]['scs'].add(s.strip())
        pt = p.get('product_type', '')
        if pt:
            companies[co]['pts'].add(pt)

    sorted_cos = sorted(companies.items(), key=lambda x: -x[1]['total'])

    for ri, (name, info) in enumerate(sorted_cos, 4):
        is_alt = ri % 2 == 0
        bg = C_ROW_ALT if is_alt else "FFFFFF"
        values = [
            name,
            info['total'],
            ', '.join(sorted(info['dts'])) if info['dts'] else '-',
            ', '.join(sorted(info['scs'])) if info['scs'] else '-',
            ', '.join(sorted(info['pts'])) if info['pts'] else '-',
        ]
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.fill   = fill(bg)
            c.border = border()
            c.font   = cell_font()
            c.alignment = Alignment(
                horizontal='center' if ci == 2 else 'left',
                vertical='center', wrap_text=True
            )
        ws.row_dimensions[ri].height = 20

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f"A3:E{len(sorted_cos)+3}"

    return ws, sorted_cos


# ── Sheet 3: Charts ─────────────────────────────────
def create_charts_sheet(wb, products, sorted_cos):
    ws = wb.create_sheet("Charts & Analysis")

    # Title
    ws.merge_cells('A1:Z1')
    c = ws['A1']
    c.value = "Thread Certified Products — Statistical Analysis"
    c.font  = Font(name='Calibri', bold=True, size=18, color=C_HEADER_FG)
    c.fill  = fill(C_TITLE_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # ── Chart 1: Device Type distribution ──
    dt_counts = Counter(
        p.get('device_type', '-') for p in products
        if p.get('device_type', '-') != '-'
    )
    dt_labels = sorted(dt_counts.keys())
    dt_vals   = [dt_counts[k] for k in dt_labels]

    DROW = 4
    ws.cell(row=DROW, column=2).value = "Device Type"
    ws.cell(row=DROW, column=3).value = "Count"
    ws.cell(row=DROW, column=2).font = Font(bold=True)
    ws.cell(row=DROW, column=3).font = Font(bold=True)
    for i, (lbl, v) in enumerate(zip(dt_labels, dt_vals)):
        ws.cell(row=DROW+1+i, column=2).value = lbl
        ws.cell(row=DROW+1+i, column=3).value = v

    chart1 = BarChart()
    chart1.type    = "col"
    chart1.title   = "Products by Device Type"
    chart1.y_axis.title = "Number of Products"
    chart1.x_axis.title = "Device Type"
    chart1.style   = 10
    chart1.width   = 20
    chart1.height  = 14
    chart1.add_data(Reference(ws, min_col=3, max_col=3, min_row=DROW, max_row=DROW+len(dt_labels)),
                    titles_from_data=True)
    chart1.set_categories(Reference(ws, min_col=2, max_col=2, min_row=DROW+1, max_row=DROW+len(dt_labels)))
    ws.add_chart(chart1, "B3")

    # ── Chart 2: Sub Category distribution ──
    sc_counts = Counter()
    for p in products:
        sc = p.get('sub_category', '-')
        if sc and sc != '-':
            for s in sc.split(', '):
                sc_counts[s.strip()] += 1
    sc_labels = sorted(sc_counts.keys())
    sc_vals   = [sc_counts[k] for k in sc_labels]

    SCROW = 4
    SCCOL = 16
    ws.cell(row=SCROW, column=SCCOL).value   = "Sub Category"
    ws.cell(row=SCROW, column=SCCOL+1).value = "Count"
    ws.cell(row=SCROW, column=SCCOL).font   = Font(bold=True)
    ws.cell(row=SCROW, column=SCCOL+1).font = Font(bold=True)
    for i, (lbl, v) in enumerate(zip(sc_labels, sc_vals)):
        ws.cell(row=SCROW+1+i, column=SCCOL).value   = lbl
        ws.cell(row=SCROW+1+i, column=SCCOL+1).value = v

    chart2 = BarChart()
    chart2.type    = "bar"   # horizontal
    chart2.title   = "Products by Sub Category"
    chart2.y_axis.title = "Sub Category"
    chart2.x_axis.title = "Count"
    chart2.style   = 10
    chart2.width   = 20
    chart2.height  = 16
    chart2.add_data(Reference(ws, min_col=SCCOL+1, max_col=SCCOL+1,
                               min_row=SCROW, max_row=SCROW+len(sc_labels)),
                    titles_from_data=True)
    chart2.set_categories(Reference(ws, min_col=SCCOL, max_col=SCCOL,
                                     min_row=SCROW+1, max_row=SCROW+len(sc_labels)))
    ws.add_chart(chart2, "O3")

    # ── Chart 3: Top 20 companies by product count ──
    CROW = 22
    ws.cell(row=CROW, column=2).value = "Company"
    ws.cell(row=CROW, column=3).value = "Products"
    ws.cell(row=CROW, column=2).font = Font(bold=True)
    ws.cell(row=CROW, column=3).font = Font(bold=True)
    top20 = sorted_cos[:20]
    for i, (name, info) in enumerate(top20):
        ws.cell(row=CROW+1+i, column=2).value = name[:40]
        ws.cell(row=CROW+1+i, column=3).value = info['total']

    chart3 = BarChart()
    chart3.type    = "bar"
    chart3.title   = "Top 20 Companies by Product Count"
    chart3.y_axis.title = "Company"
    chart3.x_axis.title = "Number of Products"
    chart3.style   = 10
    chart3.width   = 22
    chart3.height  = 16
    chart3.add_data(Reference(ws, min_col=3, max_col=3, min_row=CROW, max_row=CROW+20),
                    titles_from_data=True)
    chart3.set_categories(Reference(ws, min_col=2, max_col=2,
                                     min_row=CROW+1, max_row=CROW+20))
    ws.add_chart(chart3, "B22")

    # ── Chart 4: Product Type pie ──
    pt_counts = Counter(p.get('product_type', 'Unknown') for p in products)
    PTROW = 22
    PTCOL = 16
    ws.cell(row=PTROW, column=PTCOL).value   = "Product Type"
    ws.cell(row=PTROW, column=PTCOL+1).value = "Count"
    ws.cell(row=PTROW, column=PTCOL).font   = Font(bold=True)
    ws.cell(row=PTROW, column=PTCOL+1).font = Font(bold=True)
    for i, (lbl, v) in enumerate(pt_counts.most_common()):
        ws.cell(row=PTROW+1+i, column=PTCOL).value   = lbl
        ws.cell(row=PTROW+1+i, column=PTCOL+1).value = v

    chart4 = PieChart()
    chart4.title  = "Product Type Distribution"
    chart4.style  = 10
    chart4.width  = 16
    chart4.height = 12
    chart4.add_data(Reference(ws, min_col=PTCOL+1, max_col=PTCOL+1,
                               min_row=PTROW, max_row=PTROW+len(pt_counts)),
                    titles_from_data=True)
    chart4.set_categories(Reference(ws, min_col=PTCOL, max_col=PTCOL,
                                     min_row=PTROW+1, max_row=PTROW+len(pt_counts)))
    ws.add_chart(chart4, "O22")


# ── Main ────────────────────────────────────────────
def main():
    print(f"Reading {PRODUCTS_FILE}...")
    with open(PRODUCTS_FILE, encoding='utf-8') as f:
        data = json.load(f)

    products     = data.get('products', [])
    last_updated = data.get('last_updated', datetime.now().isoformat())
    print(f"  {len(products)} products from {len(set(p.get('company','') for p in products))} companies")

    wb = openpyxl.Workbook()

    print("Creating All Products sheet...")
    create_products_sheet(wb, products, last_updated)

    print("Creating Company Summary sheet...")
    _, sorted_cos = create_company_sheet(wb, products)

    print("Creating Charts sheet...")
    create_charts_sheet(wb, products, sorted_cos)

    print(f"Saving to {OUTPUT_PATH}...")
    wb.save(OUTPUT_PATH)
    print(f"Done! ✅  {OUTPUT_PATH}")

    # Stats
    dt_map = Counter(p.get('device_type', '-') for p in products)
    sc_map = Counter(p.get('sub_category', '-') for p in products)
    print(f"\n  DT mapped:  {sum(v for k,v in dt_map.items() if k!='-')}/{len(products)}")
    print(f"  SC mapped:  {sum(v for k,v in sc_map.items() if k!='-')}/{len(products)}")
    print(f"  Companies:  {len(sorted_cos)}")


if __name__ == '__main__':
    main()
